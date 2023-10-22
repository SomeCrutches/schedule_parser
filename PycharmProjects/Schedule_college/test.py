from icecream import ic  # type: ignore
import openpyxl  # type: ignore
import os


class ScheduleParser:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.teacher_pairs = {}
        self.tables = {}
        self.files = [f for f in os.listdir(self.folder_path) if os.path.isfile(os.path.join(self.folder_path, f))]

    def num_of_rows(self, start_row, sheet):
        current_row = 0
        number_of_rows = []
        while True:
            next_row = sheet.cell(row=start_row, column=2).value
            if next_row is None:
                number_of_rows.append(current_row)
                break
            if next_row - current_row <= 0:
                number_of_rows.append(current_row)
            current_row = next_row
            start_row += 1
        return number_of_rows

    def calculate_start_pos(self, sheet):
        row_calc = 1
        while sheet.cell(row=row_calc, column=1).value is None:
            row_calc += 1
        return row_calc + 1

    def parse_schedule(self):
        for file in self.files:
            if "~" not in file:
                self.tables[file] = []
                table_name = os.path.join(self.folder_path, file)
                workbook = openpyxl.load_workbook(table_name)
                sheet_names = workbook.sheetnames

                for name in sheet_names:
                    self.tables[file].append(name)

        for selected_table, sheets in self.tables.items():
            workbook = openpyxl.load_workbook(os.path.join(self.folder_path, selected_table))
            for sheet_selected in sheets:
                sheet = workbook[sheet_selected]
                start_pos = self.calculate_start_pos(sheet)
                rows_cnt = self.num_of_rows(start_pos, sheet)
                for columns in range(3, len(sheet_selected.split(",")) * 3 + 3, 3):
                    diff = start_pos
                    row, start_column, end_column = start_pos - 1, columns + 1, columns + 2
                    for col1 in range(start_column, end_column + 1):
                        if sheet.cell(row, col1).coordinate in sheet.merged_cells:
                            cell_value = sheet.cell(row, start_column).value
                            group = cell_value
                            break

                    for _ in range(len(rows_cnt)):
                        start_row, end_row, column = diff, diff + rows_cnt[_] - 1, 1

                        for range_ in sheet.merged_cells.ranges:
                            if range_.min_row <= start_row and range_.max_row >= end_row and range_.min_col <= column <= range_.max_col:
                                cell_value = sheet.cell(row=range_.min_row, column=range_.min_col).value
                                day_of_week = cell_value
                                break

                        day = diff
                        for row in range(day, day + rows_cnt[_]):
                            col = columns
                            cell_value = sheet.cell(row=row, column=col).value
                            if cell_value is not None:
                                pair_number = sheet.cell(row=row, column=2).value
                                auditorium = sheet.cell(row=row, column=columns).value
                                subject = sheet.cell(row=row, column=columns + 1).value
                                teacher = sheet.cell(row=row, column=columns + 2).value

                                if teacher not in self.teacher_pairs:
                                    self.teacher_pairs[teacher] = []

                                self.teacher_pairs[teacher].append(
                                    (day_of_week, pair_number, auditorium, subject, group))

                        diff += rows_cnt[_]

    def display_schedule(self, target_teacher="."):
        order_of_days = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']
        sorted_teacher_pairs = {teacher: sorted(pairs, key=lambda x: (order_of_days.index(x[0].strip().lower()), x[1]))
                                for teacher, pairs in self.teacher_pairs.items()}

        for teacher, pairs in sorted_teacher_pairs.items():
            if target_teacher in teacher:
                for pair in pairs:
                    print(
                        f"День недели: {pair[0]}, Номер пары: {pair[1]}, Аудитория: {pair[2]}, Предмет: {pair[3]}, Группа: {pair[4]}")


if __name__ == '__main__':
    folder_path = 'Tables'
    parser = ScheduleParser(folder_path)
    parser.parse_schedule()
    parser.display_schedule("Дедюхина А.А.")
